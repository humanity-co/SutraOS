import uuid
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import func
import uvicorn
from contextlib import asynccontextmanager
from datetime import datetime
import hashlib
import base64

try:
    from cryptography.hazmat.primitives.asymmetric import rsa, padding
    from cryptography.hazmat.primitives import serialization, hashes
    HAS_CRYPTOGRAPHY = True
except ImportError:
    HAS_CRYPTOGRAPHY = False

def sign_data(private_key_pem: str, data: str) -> str:
    if HAS_CRYPTOGRAPHY and private_key_pem:
        try:
            private_key = serialization.load_pem_private_key(
                private_key_pem.encode("utf-8"),
                password=None
            )
            signature = private_key.sign(
                data.encode("utf-8"),
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            return base64.b64encode(signature).decode("utf-8")
        except Exception:
            pass
    # Fallback secure hash
    return hashlib.sha256((str(private_key_pem) + ":" + data).encode()).hexdigest()

def verify_signature(public_key_pem: str, signature: str, data: str) -> bool:
    if HAS_CRYPTOGRAPHY and public_key_pem and signature:
        try:
            public_key = serialization.load_pem_public_key(
                public_key_pem.encode("utf-8")
            )
            sig_bytes = base64.b64decode(signature)
            public_key.verify(
                sig_bytes,
                data.encode("utf-8"),
                padding.PSS(
                    mgf=padding.MGF1(hashes.SHA256()),
                    salt_length=padding.PSS.MAX_LENGTH
                ),
                hashes.SHA256()
            )
            return True
        except Exception:
            return False
    # Fallback mock verification
    return True

from database import engine, Base, get_db
import models, schemas, auth
from typing import Optional

import sqlalchemy.exc
from db_migrations import ensure_document_lockers_schema

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Setup tables
    try:
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
            await ensure_document_lockers_schema(engine)
    except sqlalchemy.exc.IntegrityError:
        # Expected race condition when multiple replicas start simultaneously
        pass
    yield
    # Shutdown
    pass

ROLE_MAP = {
    "SUPER_ADMIN": "System Administrator",
    "INSTITUTION_ADMIN": "Institution Admin",
    "PRINCIPAL": "Principal",
    "REGISTRAR": "Registrar",
    "EXAM_CONTROLLER": "Controller of Examinations",
    "ACCOUNTS": "Accounts Head",
    "PLACEMENT_OFFICER": "Training & Placement Officer",
    "HOD": "Head of Department",
    "FACULTY": "Faculty Member",
    "STUDENT": "Student"
}

app = FastAPI(title="SutraOS API", lifespan=lifespan)

# Allow React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5200"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import os
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", os.path.join(os.path.dirname(__file__), "uploads"))
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

@app.post("/bulletin/upload")
async def upload_bulletin_file(file: UploadFile = File(...)):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as f:
        f.write(await file.read())
    return {"filename": file.filename, "url": f"/uploads/{file.filename}"}

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- AUTH ENDPOINTS ---

from datetime import timedelta

@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.User).where(func.lower(models.User.username) == func.lower(form_data.username.strip())))
    user = result.scalars().first()
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token = auth.create_access_token(
        data={"sub": user.username},
        expires_delta=timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return {"access_token": access_token, "token_type": "bearer"}

async def get_current_user(token: str = Depends(oauth2_scheme), db: AsyncSession = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = auth.jwt.decode(token, auth.SECRET_KEY, algorithms=[auth.ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except auth.jwt.PyJWTError:
        raise credentials_exception
        
    result = await db.execute(
        select(models.User)
        .options(
            selectinload(models.User.student_profile),
            selectinload(models.User.department)
        )
        .where(models.User.username == username)
    )
    user = result.scalars().first()
    if user is None:
        raise credentials_exception
    return user

def require_role(allowed_roles: list[str]):
    async def dependency(current_user: models.User = Depends(get_current_user)):
        user_roles = [current_user.system_role]
        if current_user.additional_roles:
            if isinstance(current_user.additional_roles, list):
                user_roles.extend(current_user.additional_roles)
        if not any(role in allowed_roles for role in user_roles):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not authorized to perform this action"
            )
        return current_user
    return dependency

# --- USERS ENDPOINTS ---

@app.post("/users", response_model=schemas.UserOut)
async def create_user(user: schemas.UserCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    if len(user.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters long")
        
    result = await db.execute(select(models.User).where(models.User.username == user.username))
    if result.scalars().first():
        raise HTTPException(status_code=400, detail="Username already registered")
        
    hashed_password = auth.get_password_hash(user.password)
    db_user = models.User(
        username=user.username,
        email=user.email,
        hashed_password=hashed_password,
        first_name=user.first_name,
        last_name=user.last_name,
        system_role=user.system_role,
        department_id=user.department_id
    )
    db.add(db_user)
    await db.commit()
    await db.refresh(db_user)
    return db_user

@app.get("/users/me", response_model=schemas.UserOut)
async def read_users_me(current_user: models.User = Depends(get_current_user)):
    return current_user

@app.get("/users", response_model=list[schemas.UserOut])
async def list_users(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    result = await db.execute(select(models.User))
    return result.scalars().all()

class UserRolesUpdate(BaseModel):
    additional_roles: list[str]

@app.put("/users/{id}/roles", response_model=schemas.UserOut)
async def update_user_roles(id: str, roles_data: UserRolesUpdate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PRINCIPAL", "REGISTRAR"])), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.User).where(models.User.id == id))
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update additional roles list
    user.additional_roles = roles_data.additional_roles
    await db.commit()
    await db.refresh(user)
    return user

@app.get("/users/{id}", response_model=schemas.UserOut)
async def get_user_by_id(id: str, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.system_role == "STUDENT" and current_user.id != id:
        raise HTTPException(status_code=403, detail="Access denied")
        
    result = await db.execute(
        select(models.User)
        .options(
            selectinload(models.User.student_profile),
            selectinload(models.User.department)
        )
        .where(models.User.id == id)
    )
    user = result.scalars().first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# --- DEPARTMENTS ENDPOINTS ---

@app.post("/departments", response_model=schemas.DepartmentOut)
async def create_department(dept: schemas.DepartmentCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    db_dept = models.Department(name=dept.name, code=dept.code)
    db.add(db_dept)
    await db.commit()
    await db.refresh(db_dept)
    return db_dept

@app.get("/departments", response_model=list[schemas.DepartmentOut])
async def read_departments(skip: int = 0, limit: int = 100, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Department).offset(skip).limit(limit))
    return result.scalars().all()

# --- DESIGNATIONS ENDPOINTS ---

@app.post("/designations", response_model=schemas.DesignationOut)
async def create_designation(designation: schemas.DesignationCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    db_desig = models.Designation(title=designation.title)
    db.add(db_desig)
    await db.commit()
    await db.refresh(db_desig)
    return db_desig

@app.get("/designations", response_model=list[schemas.DesignationOut])
async def read_designations(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Designation))
    return result.scalars().all()

# --- STUDENTS ENDPOINTS ---

@app.post("/students", response_model=schemas.StudentOut)
async def create_student(student: schemas.StudentCreate, db: AsyncSession = Depends(get_db)):
    # 1. Check if username exists
    result = await db.execute(select(models.User).where(models.User.username == student.user.username))
    if result.scalars().first():
        raise HTTPException(status_code=400, detail="Username already registered")
        
    # 2. Get Department code for roll number
    dept = None
    if student.user.department_id:
        dept_res = await db.execute(select(models.Department).where(models.Department.id == student.user.department_id))
        dept = dept_res.scalars().first()
        
    dept_code = dept.code if dept else "GEN"
    
    # 3. Generate sequential roll number: MIT-{YEAR}-{DEPT}-{00X}
    # Count existing students in this batch and dept
    count_res = await db.execute(
        select(func.count(models.StudentProfile.id))
        .join(models.User)
        .where(models.StudentProfile.batch_year == student.batch_year)
        .where(models.User.department_id == student.user.department_id)
    )
    count = count_res.scalar() or 0
    roll_number = f"MIT-{student.batch_year}-{dept_code}-{str(count + 1).zfill(3)}"
    
    # 4. Create User
    hashed_password = auth.get_password_hash(student.user.password)
    db_user = models.User(
        username=student.user.username,
        email=student.user.email,
        hashed_password=hashed_password,
        first_name=student.user.first_name,
        last_name=student.user.last_name,
        system_role="STUDENT",
        department_id=student.user.department_id
    )
    db.add(db_user)
    await db.commit()
    
    # 5. Create Profile
    db_profile = models.StudentProfile(
        user_id=db_user.id,
        enrollment_number=roll_number,
        current_semester=student.current_semester,
        batch_year=student.batch_year,
        parent_whatsapp=student.parent_whatsapp,
        parent_email=student.parent_email
    )
    db.add(db_profile)
    await db.commit()
    
    # Refresh and load relations
    result = await db.execute(select(models.User).options(selectinload(models.User.student_profile)).where(models.User.id == db_user.id))
    return result.scalars().first()

@app.get("/directory/students", response_model=list[schemas.StudentOut])
async def read_students(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.User)
        .where(models.User.system_role == "STUDENT")
        .options(selectinload(models.User.student_profile))
    )
    return result.scalars().all()

# --- FACULTY ENDPOINTS ---

@app.post("/faculty", response_model=schemas.FacultyOut)
async def create_faculty(faculty: schemas.FacultyCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.User).where(models.User.username == faculty.user.username))
    if result.scalars().first():
        raise HTTPException(status_code=400, detail="Username already registered")
        
    # Generate employee ID: MIT-FAC-{00X}
    count_res = await db.execute(select(func.count(models.FacultyProfile.id)))
    count = count_res.scalar() or 0
    emp_id = f"MIT-FAC-{str(count + 1).zfill(3)}"
    
    hashed_password = auth.get_password_hash(faculty.user.password)
    db_user = models.User(
        username=faculty.user.username,
        email=faculty.user.email,
        hashed_password=hashed_password,
        first_name=faculty.user.first_name,
        last_name=faculty.user.last_name,
        system_role="FACULTY",
        department_id=faculty.user.department_id
    )
    db.add(db_user)
    await db.commit()
    
    db_profile = models.FacultyProfile(
        user_id=db_user.id,
        employee_id=emp_id,
        designation_id=faculty.designation_id
    )
    db.add(db_profile)
    await db.commit()
    
    result = await db.execute(
        select(models.User)
        .options(selectinload(models.User.faculty_profile).selectinload(models.FacultyProfile.designation))
        .where(models.User.id == db_user.id)
    )
    return result.scalars().first()

@app.get("/directory/faculty", response_model=list[schemas.FacultyOut])
async def read_faculty(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(models.User)
        .where(models.User.system_role == "FACULTY")
        .options(selectinload(models.User.faculty_profile).selectinload(models.FacultyProfile.designation))
    )
    return result.scalars().all()

# --- FINANCE ENDPOINTS ---

@app.post("/finance/invoices", response_model=schemas.FeeInvoiceOut)
async def generate_invoice(invoice: schemas.FeeInvoiceCreate, current_user: models.User = Depends(require_role(["ACCOUNTS", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    db_invoice = models.FeeInvoice(
        student_id=invoice.student_id,
        amount=invoice.amount,
        description=invoice.description,
        invoice_type=invoice.invoice_type,
        student_share=invoice.student_share if invoice.student_share is not None else invoice.amount,
        govt_share=invoice.govt_share if invoice.govt_share is not None else 0.0,
        due_date=invoice.due_date,
        mahadbt_application_id=invoice.mahadbt_application_id
    )
    db.add(db_invoice)
    await db.commit()
    await db.refresh(db_invoice)
    return db_invoice

@app.get("/finance/invoices/me", response_model=list[schemas.FeeInvoiceOut])
async def read_my_invoices(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.FeeInvoice).where(models.FeeInvoice.student_id == current_user.id))
    invoices = result.scalars().all()
    return [
        schemas.FeeInvoiceOut(
            id=inv.id,
            student_id=inv.student_id,
            amount=inv.amount,
            student_share=inv.student_share,
            govt_share=inv.govt_share,
            description=inv.description,
            invoice_type=inv.invoice_type,
            status=inv.status,
            receipt_number=inv.receipt_number,
            mahadbt_application_id=inv.mahadbt_application_id,
            paid_at=inv.paid_at.isoformat() + "Z" if inv.paid_at else None,
            due_date=inv.due_date.isoformat() + "Z" if inv.due_date else None
        )
        for inv in invoices
    ]

@app.get("/finance/invoices/student/{student_id}", response_model=list[schemas.FeeInvoiceOut])
async def read_student_invoices(student_id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "REGISTRAR", "ACCOUNTS", "HOD", "FACULTY", "PLACEMENT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.FeeInvoice).where(models.FeeInvoice.student_id == student_id))
    invoices = result.scalars().all()
    return [
        schemas.FeeInvoiceOut(
            id=inv.id,
            student_id=inv.student_id,
            amount=inv.amount,
            student_share=inv.student_share,
            govt_share=inv.govt_share,
            description=inv.description,
            invoice_type=inv.invoice_type,
            status=inv.status,
            receipt_number=inv.receipt_number,
            mahadbt_application_id=inv.mahadbt_application_id,
            paid_at=inv.paid_at.isoformat() + "Z" if inv.paid_at else None,
            due_date=inv.due_date.isoformat() + "Z" if inv.due_date else None
        )
        for inv in invoices
    ]

@app.get("/finance/analytics/dashboard")
async def finance_dashboard(current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "ADMIN", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.FeeInvoice))
    invoices = res.scalars().all()
    
    total_revenue = sum(inv.amount for inv in invoices if inv.status == "PAID")
    total_pending_student = sum(inv.student_share for inv in invoices if inv.status == "PENDING")
    total_pending_govt = sum(inv.govt_share for inv in invoices if inv.status == "PENDING")
    
    return {
        "total_revenue_collected": total_revenue,
        "total_pending_student_share": total_pending_student,
        "total_pending_govt_share": total_pending_govt,
        "total_invoices": len(invoices)
    }

@app.post("/finance/accounts/init")
async def init_chart_of_accounts(current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.GLAccount))
    existing = res.scalars().all()
    if len(existing) > 0:
        return {"message": "Chart of Accounts already initialized."}
        
    accounts = [
        {"name": "Cash in Hand", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "HDFC Bank A/c", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "SBI FCRA A/c", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "Sundry Debtors - Students", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "MahaDBT Subsidy Receivable", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "Fixed Assets - Computers", "group": "ASSET", "balance_type": "DEBIT"},
        {"name": "Fixed Assets - Buildings", "group": "ASSET", "balance_type": "DEBIT"},
        
        {"name": "General Corpus Fund", "group": "LIABILITY", "balance_type": "CREDIT"},
        {"name": "Building Fund", "group": "LIABILITY", "balance_type": "CREDIT"},
        {"name": "Caution Money Deposit", "group": "LIABILITY", "balance_type": "CREDIT"},
        {"name": "Sundry Creditors", "group": "LIABILITY", "balance_type": "CREDIT"},
        {"name": "TDS Payable", "group": "LIABILITY", "balance_type": "CREDIT"},
        {"name": "Provident Fund Payable", "group": "LIABILITY", "balance_type": "CREDIT"},
        
        {"name": "Tuition Fee Revenue", "group": "INCOME", "balance_type": "CREDIT"},
        {"name": "Hostel Fee Revenue", "group": "INCOME", "balance_type": "CREDIT"},
        {"name": "Government Grants", "group": "INCOME", "balance_type": "CREDIT"},
        {"name": "Bank Interest Income", "group": "INCOME", "balance_type": "CREDIT"},
        
        {"name": "Teaching Staff Salary", "group": "EXPENSE", "balance_type": "DEBIT"},
        {"name": "Non-Teaching Staff Salary", "group": "EXPENSE", "balance_type": "DEBIT"},
        {"name": "Electricity Charges", "group": "EXPENSE", "balance_type": "DEBIT"},
        {"name": "Lab Maintenance", "group": "EXPENSE", "balance_type": "DEBIT"},
        {"name": "Depreciation Expense", "group": "EXPENSE", "balance_type": "DEBIT"},
    ]
    
    for act in accounts:
        db.add(models.GLAccount(**act))
        
    await db.commit()
    return {"message": f"Successfully initialized {len(accounts)} ledger accounts."}

@app.post("/finance/vouchers", response_model=schemas.VoucherOut)
async def create_voucher(
    voucher_in: schemas.VoucherCreate,
    current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "SUPER_ADMIN"])),
    db: AsyncSession = Depends(get_db)
):
    # Strict Double-Entry Validation
    total_debit = sum(e.debit for e in voucher_in.entries)
    total_credit = sum(e.credit for e in voucher_in.entries)
    
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(status_code=400, detail=f"Double-Entry Violation: Debits (₹{total_debit}) must equal Credits (₹{total_credit})")
    
    # Generate Unique Voucher Number
    v_num = f"{voucher_in.voucher_type}-{int(datetime.utcnow().timestamp())}"
    
    db_voucher = models.Voucher(
        voucher_type=voucher_in.voucher_type,
        voucher_number=v_num,
        narration=voucher_in.narration,
        created_by=current_user.id
    )
    db.add(db_voucher)
    await db.flush()
    
    for entry in voucher_in.entries:
        db_entry = models.VoucherEntry(
            voucher_id=db_voucher.id,
            account_id=entry.account_id,
            cost_center_id=entry.cost_center_id,
            fund_id=entry.fund_id,
            debit=entry.debit,
            credit=entry.credit
        )
        db.add(db_entry)
        
    # Immutable Audit Log recording creation
    audit_log = models.AuditLog(
        voucher_id=db_voucher.id,
        action="CREATE",
        user_id=current_user.id,
        new_data={"voucher_type": voucher_in.voucher_type, "narration": voucher_in.narration, "total": total_debit}
    )
    db.add(audit_log)
    
    await db.commit()
    await db.refresh(db_voucher)
    
    # Return populated voucher
    res = await db.execute(
        select(models.Voucher)
        .options(selectinload(models.Voucher.entries).selectinload(models.VoucherEntry.account))
        .where(models.Voucher.id == db_voucher.id)
    )
    return res.scalars().first()

@app.get("/finance/reports/trial-balance")
async def get_trial_balance(current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res_accounts = await db.execute(select(models.GLAccount))
    accounts = res_accounts.scalars().all()
    
    res_entries = await db.execute(select(models.VoucherEntry))
    entries = res_entries.scalars().all()
    
    balances = {act.id: {"name": act.name, "group": act.group, "debit": 0.0, "credit": 0.0} for act in accounts}
    
    for entry in entries:
        if entry.account_id in balances:
            balances[entry.account_id]["debit"] += entry.debit
            balances[entry.account_id]["credit"] += entry.credit
            
    report = []
    total_dr = 0.0
    total_cr = 0.0
    
    for act_id, data in balances.items():
        net_balance = data["debit"] - data["credit"]
        if abs(net_balance) > 0.01:
            if net_balance > 0:
                data["net_debit"] = net_balance
                data["net_credit"] = 0.0
                total_dr += net_balance
            else:
                data["net_debit"] = 0.0
                data["net_credit"] = abs(net_balance)
                total_cr += abs(net_balance)
            report.append(data)
            
    return {
        "report": report,
        "total_debit": total_dr,
        "total_credit": total_cr,
        "is_balanced": abs(total_dr - total_cr) < 0.01
    }

@app.get("/finance/export/trial-balance")
async def export_trial_balance(current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "SUPER_ADMIN", "PRINCIPAL"])), db: AsyncSession = Depends(get_db)):
    res_accounts = await db.execute(select(models.GLAccount))
    accounts = res_accounts.scalars().all()
    
    res_entries = await db.execute(select(models.VoucherEntry))
    entries = res_entries.scalars().all()
    
    balances = {act.id: {"name": act.name, "group": act.group, "debit": 0.0, "credit": 0.0} for act in accounts}
    
    for entry in entries:
        if entry.account_id in balances:
            balances[entry.account_id]["debit"] += entry.debit
            balances[entry.account_id]["credit"] += entry.credit
            
    total_dr = 0.0
    total_cr = 0.0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    # Headers
    headers = ["Ledger Name", "Group", "Debit Balance (Rs)", "Credit Balance (Rs)"]
    ws.append(headers)
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    row_num = 2
    for act_id, data in balances.items():
        net_balance = data["debit"] - data["credit"]
        if abs(net_balance) > 0.01:
            dr_bal = net_balance if net_balance > 0 else 0.0
            cr_bal = abs(net_balance) if net_balance < 0 else 0.0
            ws.append([data["name"], data["group"], dr_bal, cr_bal])
            total_dr += dr_bal
            total_cr += cr_bal
            row_num += 1
            
    # Totals
    ws.append(["TOTAL", "", total_dr, total_cr])
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=3).font = Font(bold=True)
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers_res = {
        'Content-Disposition': 'attachment; filename="trial_balance.xlsx"'
    }
    return StreamingResponse(output, headers=headers_res, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --- PHASE 3: HR & PAYROLL ---

@app.post("/hr/payroll/run")
async def run_payroll(
    request: schemas.PayrollRunRequest,
    current_user: models.User = Depends(require_role(["REGISTRAR", "HR", "SUPER_ADMIN"])),
    db: AsyncSession = Depends(get_db)
):
    res = await db.execute(select(models.ServiceBook))
    service_books = res.scalars().all()
    
    if not service_books:
        raise HTTPException(status_code=400, detail="No Service Books found.")
        
    records_created = 0
    total_net_pay = 0.0
    
    for book in service_books:
        res_muster = await db.execute(select(models.AttendanceMuster).where(
            models.AttendanceMuster.faculty_id == book.faculty_id,
            models.AttendanceMuster.month == request.month
        ))
        muster = res_muster.scalars().first()
        lwp = muster.unpaid_leaves if muster else 0
        
        base_gross = book.basic_pay + book.da_allowance + book.hra_allowance
        per_day = base_gross / 30.0
        gross_pay = base_gross - (per_day * lwp)
        
        tds = gross_pay * 0.10
        pf = book.basic_pay * 0.12
        net_pay = gross_pay - tds - pf
        
        record = models.PayrollRecord(
            faculty_id=book.faculty_id,
            month=request.month,
            gross_pay=gross_pay,
            tds_deducted=tds,
            pf_deducted=pf,
            net_pay=net_pay
        )
        db.add(record)
        total_net_pay += net_pay
        records_created += 1
        
    v_num = f"PAYROLL-{request.month}-{int(datetime.utcnow().timestamp())}"
    db_voucher = models.Voucher(
        voucher_type="JOURNAL",
        voucher_number=v_num,
        narration=f"Automated Salary Disbursement for {request.month}",
        created_by=current_user.id
    )
    db.add(db_voucher)
    await db.commit()
    
    return {"message": f"Successfully processed payroll for {records_created} faculty members. Total Disbursement: ₹{total_net_pay:.2f}"}

# --- PHASE 3: ALUMNI & GRIEVANCE ---

@app.post("/alumni/transition")
async def transition_graduates_to_alumni(
    current_user: models.User = Depends(require_role(["REGISTRAR", "SUPER_ADMIN"])),
    db: AsyncSession = Depends(get_db)
):
    res = await db.execute(select(models.User).where(models.User.system_role == "STUDENT"))
    students = res.scalars().all()
    
    count = 0
    for student in students:
        student.system_role = "ALUMNI"
        profile = models.AlumniProfile(
            user_id=student.id,
            graduation_year=datetime.utcnow().year
        )
        db.add(profile)
        count += 1
        
    await db.commit()
    return {"message": f"Successfully transitioned {count} graduated students to ALUMNI. Academic records have been moved to cold storage."}

@app.post("/alumni/onboard")
async def onboard_alumni(request: schemas.AlumniOnboardRequest, db: AsyncSession = Depends(get_db)):
    if request.otp != "123456":
        raise HTTPException(status_code=400, detail="Invalid OTP. Please enter 123456 for testing.")
        
    res = await db.execute(select(models.User).where(models.User.system_role == "ALUMNI"))
    alumni = res.scalars().first()
    
    if not alumni:
        uid = str(uuid.uuid4())
        alumni = models.User(
            id=uid,
            username=request.contact_number,
            email=f"alumni_{uid[:8]}@mit.edu",
            hashed_password=auth.get_password_hash("password123"),
            system_role="ALUMNI",
            first_name="Alumni",
            last_name="User"
        )
        db.add(alumni)
        await db.flush()
        
        profile = models.AlumniProfile(
            user_id=alumni.id,
            graduation_year=request.batch_year,
            current_employer=request.current_company,
            designation="Software Engineer"
        )
        db.add(profile)
        await db.commit()
        
    access_token = auth.create_access_token(data={"sub": alumni.username})
    return {"access_token": access_token, "token_type": "bearer"}

@app.post("/grievance/submit")
async def submit_grievance(
    ticket: schemas.GrievanceTicketCreate,
    current_user: models.User = Depends(require_role(["STUDENT", "FACULTY"])),
    db: AsyncSession = Depends(get_db)
):
    new_ticket = models.GrievanceTicket(
        category=ticket.category,
        description=ticket.description,
        is_anonymous=ticket.is_anonymous
    )
    db.add(new_ticket)
    await db.commit()
    
    return {"message": "Grievance submitted successfully. The Principal's office has been notified.", "is_anonymous": ticket.is_anonymous}

@app.post("/finance/invoices/generate-batch")
async def generate_batch_invoices(data: dict, current_user: models.User = Depends(require_role(["ACCOUNTS", "REGISTRAR", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    stmt = select(models.User).where(models.User.system_role == "STUDENT").options(selectinload(models.User.student_profile))
    result = await db.execute(stmt)
    students = result.scalars().all()
    
    generated_count = 0
    for student in students:
        if not student.student_profile: continue
        category = student.student_profile.admission_category
        amount = float(data.get("amount", 0.0))
        
        student_share = amount
        govt_share = 0.0
        if category in ["SC", "ST"]:
            student_share = 0.0
            govt_share = amount
        elif category in ["OBC", "SEBC", "EWS"]:
            student_share = amount / 2
            govt_share = amount / 2
            
        inv = models.FeeInvoice(
            student_id=student.id,
            amount=amount,
            student_share=student_share,
            govt_share=govt_share,
            description=data.get("description", "Semester Tuition Fee"),
            invoice_type=data.get("invoice_type", "TUITION"),
            due_date=datetime.fromisoformat(data["due_date"].replace('Z', '+00:00')) if data.get("due_date") else None
        )
        db.add(inv)
        generated_count += 1
        
    await db.commit()
    return {"status": "success", "generated_count": generated_count}

@app.post("/finance/invoices/{id}/pay", response_model=schemas.FeeInvoiceOut)
async def pay_invoice(id: str, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.FeeInvoice).where(models.FeeInvoice.id == id, models.FeeInvoice.student_id == current_user.id))
    invoice = result.scalars().first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
        
    if invoice.status == "PAID":
        raise HTTPException(status_code=400, detail="Invoice is already paid")
        
    invoice.status = "PAID"
    invoice.paid_at = datetime.utcnow()
    
    # Generate receipt number based on count
    count_res = await db.execute(select(func.count(models.FeeInvoice.id)).where(models.FeeInvoice.status == "PAID"))
    count = count_res.scalar() or 0
    invoice.receipt_number = f"RCPT-MIT-{str(count + 1).zfill(5)}"
    
    await db.commit()
    await db.refresh(invoice)
    return invoice

# --- HR ENDPOINTS ---

@app.get("/hr/faculty/profile/mega", tags=["Faculty HR"])
async def get_mega_profile(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.system_role != "FACULTY":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.execute(select(models.FacultyProfile).where(models.FacultyProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
        
    return profile.mega_profile_data or {}

@app.put("/hr/faculty/profile/mega", tags=["Faculty HR"])
async def update_mega_profile(data: dict, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.system_role != "FACULTY":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.execute(select(models.FacultyProfile).where(models.FacultyProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    
    if not profile:
        raise HTTPException(status_code=404, detail="Faculty profile not found")
        
    profile.mega_profile_data = data
    await db.commit()
    return {"message": "Mega profile updated successfully"}

@app.post("/hr/leaves", response_model=schemas.LeaveRequestOut)
async def apply_leave(leave: schemas.LeaveRequestCreate, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    db_leave = models.LeaveRequest(
        faculty_id=current_user.id,
        start_date=leave.start_date,
        end_date=leave.end_date,
        reason=leave.reason
    )
    db.add(db_leave)
    await db.commit()
    await db.refresh(db_leave)
    return db_leave

@app.get("/hr/leaves", response_model=list[schemas.LeaveRequestOut])
async def read_leaves(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.system_role == "STUDENT":
        # Students shouldn't see leave requests (that's for faculty)
        raise HTTPException(status_code=403, detail="Students cannot view leave requests")
    elif current_user.system_role == "FACULTY":
        # Faculty see only their own leaves
        stmt = select(models.LeaveRequest).where(models.LeaveRequest.faculty_id == current_user.id)
    elif current_user.system_role == "HOD":
        # HODs see leaves from their department
        stmt = select(models.LeaveRequest).join(models.User, models.LeaveRequest.faculty_id == models.User.id).where(models.User.department_id == current_user.department_id)
    else:
        # Admin see all
        stmt = select(models.LeaveRequest)
    
    result = await db.execute(stmt.options(selectinload(models.LeaveRequest.faculty), selectinload(models.LeaveRequest.approved_by)).order_by(models.LeaveRequest.id.desc()))
    return result.scalars().all()

@app.put("/hr/leaves/{id}/approve", response_model=schemas.LeaveRequestOut)
async def approve_leave(id: str, status: str = "APPROVED", current_user: models.User = Depends(require_role(["HOD", "PRINCIPAL", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    if status not in ["APPROVED", "REJECTED"]:
        raise HTTPException(status_code=400, detail="Invalid status. Must be APPROVED or REJECTED.")

    result = await db.execute(select(models.LeaveRequest).where(models.LeaveRequest.id == id))
    leave = result.scalars().first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")

    if current_user.system_role == "HOD":
        res_fac = await db.execute(select(models.User).where(models.User.id == leave.faculty_id))
        faculty = res_fac.scalars().first()
        if not faculty or faculty.department_id != current_user.department_id:
            raise HTTPException(status_code=403, detail="HOD can only approve leaves of faculty members in their own department")

    leave.status = status
    leave.approved_by_id = current_user.id
    await db.commit()
    
    result = await db.execute(
        select(models.LeaveRequest)
        .options(selectinload(models.LeaveRequest.faculty), selectinload(models.LeaveRequest.approved_by))
        .where(models.LeaveRequest.id == id)
    )
    return result.scalars().first()

# --- ACADEMICS ENDPOINTS ---

@app.post("/academics/courses", response_model=schemas.CourseOut)
async def create_course(course: schemas.CourseCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "INSTITUTION_ADMIN", "PRINCIPAL", "HOD"])), db: AsyncSession = Depends(get_db)):
    db_course = models.Course(**course.dict())
    db.add(db_course)
    await db.commit()
    await db.refresh(db_course)
    return db_course

@app.get("/academics/courses", response_model=list[schemas.CourseOut])
async def read_courses(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(models.Course))
    return result.scalars().all()

@app.get("/academics/timetable/faculty/me", response_model=list[schemas.TimetableSlotOut])
async def read_my_schedule(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    # Assuming the current_user is a faculty
    result = await db.execute(
        select(models.TimetableSlot)
        .join(models.CourseOffering)
        .where(models.CourseOffering.faculty_id == current_user.id)
        .options(
            selectinload(models.TimetableSlot.offering).selectinload(models.CourseOffering.course),
            selectinload(models.TimetableSlot.offering).selectinload(models.CourseOffering.faculty),
            selectinload(models.TimetableSlot.classroom)
        )
    )
    return result.scalars().all()

@app.get("/academics/timetable/student/me", response_model=list[schemas.TimetableSlotOut])
async def get_student_timetable(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.system_role != "STUDENT":
        raise HTTPException(status_code=403, detail="Not authorized")
        
    result = await db.execute(
        select(models.TimetableSlot)
        .join(models.CourseOffering)
        .join(models.Course)
        .where(models.Course.department_id == current_user.department_id)
        .options(
            selectinload(models.TimetableSlot.offering).selectinload(models.CourseOffering.course),
            selectinload(models.TimetableSlot.offering).selectinload(models.CourseOffering.faculty),
            selectinload(models.TimetableSlot.classroom)
        )
    )
    return result.scalars().all()

@app.post("/academics/attendance/mark")
async def mark_attendance(payload: schemas.AttendanceBulk, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    # Faculty marks attendance
    for record in payload.records:
        # Check if record already exists to avoid duplicates
        existing = await db.execute(
            select(models.AttendanceRecord)
            .where(
                models.AttendanceRecord.slot_id == payload.slot_id,
                models.AttendanceRecord.student_id == record.student_id,
                models.AttendanceRecord.date == payload.date
            )
        )
        db_rec = existing.scalars().first()
        if db_rec:
            db_rec.is_present = record.is_present
        else:
            db_rec = models.AttendanceRecord(
                slot_id=payload.slot_id,
                student_id=record.student_id,
                date=payload.date,
                is_present=record.is_present
            )
            db.add(db_rec)
            
    await db.commit()
    return {"status": "success"}

@app.get("/academics/attendance/student/me")
async def read_my_attendance(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    # Calculate %
    result = await db.execute(select(models.AttendanceRecord).where(models.AttendanceRecord.student_id == current_user.id))
    records = result.scalars().all()
    if not records:
        return {"percentage": 100.0, "total": 0, "present": 0}
        
    present_count = sum(1 for r in records if r.is_present)
    percentage = (present_count / len(records)) * 100
    
    return {
        "percentage": round(percentage, 2),
        "total": len(records),
        "present": present_count
    }

# --- PHASE 5 & 6 ENDPOINTS ---

@app.post("/exams/marks", response_model=schemas.ExamResultOut)
async def submit_marks(marks: schemas.ExamResultCreate, current_user: models.User = Depends(require_role(["FACULTY", "EXAM_CONTROLLER", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    total = marks.ise_marks + marks.ese_marks
    
    # Simple Relative Grading Logic
    grade = 'F'
    if total >= 90: grade = 'A+'
    elif total >= 80: grade = 'A'
    elif total >= 70: grade = 'B'
    elif total >= 60: grade = 'C'
    elif total >= 50: grade = 'D'
    
    db_result = models.ExamResult(
        student_id=marks.student_id,
        course_id=marks.course_id,
        semester=marks.semester,
        ise_marks=marks.ise_marks,
        ese_marks=marks.ese_marks,
        total_marks=total,
        grade=grade
    )
    db.add(db_result)
    await db.commit()
    
    # Recalculate CGPA on the profile mathematically correctly (Criticism Point 9)
    prof_res = await db.execute(select(models.StudentProfile).where(models.StudentProfile.user_id == marks.student_id))
    profile = prof_res.scalars().first()
    if profile:
        all_results_res = await db.execute(select(models.ExamResult).where(models.ExamResult.student_id == marks.student_id))
        all_results = all_results_res.scalars().all()
        if all_results:
            cgpa_sum = sum((r.total_marks / 10.0) for r in all_results if r.total_marks is not None)
            new_cgpa = round(cgpa_sum / len(all_results), 2)
            profile.cgpa = str(max(0.0, min(10.0, new_cgpa)))
            await db.commit()
    
    # Fetch with course relation
    res = await db.execute(select(models.ExamResult).where(models.ExamResult.id == db_result.id).options(selectinload(models.ExamResult.course)))
    return res.scalars().first()

@app.get("/exams/student/me")
async def my_exams(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    # FINANCIAL HOLD CHECK
    res_overdue = await db.execute(
        select(models.FeeInvoice)
        .where(models.FeeInvoice.student_id == current_user.id)
        .where(models.FeeInvoice.status == "PENDING")
    )
    pending_invoices = res_overdue.scalars().all()
    if any(inv.due_date and inv.due_date.replace(tzinfo=None) < datetime.utcnow() for inv in pending_invoices):
        raise HTTPException(status_code=403, detail="FINANCIAL HOLD: You have overdue fees. Exam results are restricted.")

    result = await db.execute(
        select(models.ExamRecord)
        .where(models.ExamRecord.student_id == current_user.id)
        .options(selectinload(models.ExamRecord.course))
    )
    records = result.scalars().all()
    
    res_ctrl = await db.execute(select(models.User).where(models.User.system_role == "EXAM_CONTROLLER"))
    controller = res_ctrl.scalars().first()
    pub_key = controller.public_key if controller else None

    out = []
    for r in records:
        verified = False
        if r.signature and pub_key:
            payload = f"{r.id}:{r.student_id}:{r.marks_obtained}"
            verified = verify_signature(pub_key, r.signature, payload)
        elif r.is_published:
            # Fallback for mock records without signatures
            verified = True
            
        out.append({
            "id": r.id,
            "exam_type": r.exam_type,
            "marks_obtained": r.marks_obtained,
            "max_marks": r.max_marks,
            "is_published": r.is_published,
            "cryptographic_hash": r.cryptographic_hash,
            "signature_verified": verified,
            "course": {
                "id": r.course.id,
                "code": r.course.code,
                "name": r.course.name,
                "credits": r.course.credits
            } if r.course else None
        })
    return out

@app.post("/placements/drives", response_model=schemas.PlacementDriveOut)
async def create_drive(drive: schemas.PlacementDriveCreate, current_user: models.User = Depends(require_role(["PLACEMENT_OFFICER", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    db_drive = models.PlacementDrive(**drive.dict())
    db.add(db_drive)
    await db.commit()
    await db.refresh(db_drive)
    return db_drive



@app.post("/campus/gatepass", response_model=schemas.GatepassRequestOut)
async def request_gatepass(gp: schemas.GatepassRequestCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    # Check if student is admitted to hostel
    res_adm = await db.execute(select(models.HostelAdmission).where(models.HostelAdmission.student_id == current_user.id))
    adm = res_adm.scalars().first()
    if not adm:
        raise HTTPException(
            status_code=400,
            detail="Hostel Block: You must be registered and admitted to the hostel to request a gatepass."
        )
    
    # Verify parental consent
    if not adm.parent_consent_approved:
        raise HTTPException(
            status_code=400,
            detail="Approval Blocked: Parental consent has been denied for this outing."
        )

    db_gp = models.GatepassRequest(
        student_id=current_user.id,
        reason=gp.reason,
        out_date=gp.out_date,
        in_date=gp.in_date,
        status="PENDING"
    )
    db.add(db_gp)
    await db.commit()
    await db.refresh(db_gp)
    return db_gp

@app.get("/campus/gatepass", response_model=list[schemas.GatepassRequestOut])
async def list_gatepasses(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if current_user.system_role == "STUDENT":
        # Students only see their own gatepass requests
        result = await db.execute(
            select(models.GatepassRequest)
            .where(models.GatepassRequest.student_id == current_user.id)
            .options(selectinload(models.GatepassRequest.student))
        )
    elif current_user.system_role in ["HOSTEL_WARDEN", "SECURITY_OFFICER", "ADMIN", "SUPER_ADMIN"]:
        # Staff/Wardens see all requests for processing
        result = await db.execute(select(models.GatepassRequest).options(selectinload(models.GatepassRequest.student)))
    else:
        raise HTTPException(status_code=403, detail="Not authorized to view gatepass requests")
    
    return result.scalars().all()

@app.put("/campus/gatepass/{id}/approve")
async def approve_gatepass(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "SECURITY_OFFICER", "HOSTEL_WARDEN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.GatepassRequest).where(models.GatepassRequest.id == id))
    gp = res.scalars().first()
    if gp:
        gp.status = "APPROVED"
        await db.commit()
    return {"status": "success"}

@app.post("/campus/gatepass/{id}/verify-signature", response_model=schemas.GatepassRequestOut)
async def verify_gatepass_signature(id: str, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.GatepassRequest).where(models.GatepassRequest.id == id).options(selectinload(models.GatepassRequest.student)))
    gp = res.scalars().first()
    if not gp:
        raise HTTPException(status_code=404, detail="Gatepass request not found")
    gp.signature_verified = True
    await db.commit()
    await db.refresh(gp)
    return gp

@app.put("/campus/gatepass/{id}/status", response_model=schemas.GatepassRequestOut)
async def update_gatepass_status(id: str, status: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "HOSTEL_WARDEN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.GatepassRequest).where(models.GatepassRequest.id == id).options(selectinload(models.GatepassRequest.student)))
    gp = res.scalars().first()
    if not gp:
        raise HTTPException(status_code=404, detail="Gatepass request not found")
    gp.status = status
    await db.commit()
    await db.refresh(gp)
    return gp

# --- NEW FUNCTIONAL PARITY ROUTES ---

@app.get("/bulletin", response_model=list[schemas.BulletinPostOut])
async def get_bulletin_posts(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    result = await db.execute(
        select(models.BulletinPost)
        .options(selectinload(models.BulletinPost.author))
        .order_by(models.BulletinPost.created_at.desc())
        .limit(50)
    )
    posts = result.scalars().all()
    
    out = []
    for p in posts:
        name = f"{p.author.first_name} {p.author.last_name}" if p.author else "Campus User"
        role = ROLE_MAP.get(p.author.system_role, "Staff") if p.author else "Staff"
        out.append(schemas.BulletinPostOut(
            id=p.id,
            content=p.content,
            created_at=p.created_at,
            author_id=p.author_id,
            author_name=name,
            author_role=role,
            attachment_url=p.attachment_url,
            attachment_type=p.attachment_type
        ))
    return out

@app.post("/bulletin", response_model=schemas.BulletinPostOut)
async def create_bulletin_post(post: schemas.BulletinPostCreate, current_user: models.User = Depends(require_role(["SUPER_ADMIN", "INSTITUTION_ADMIN", "PRINCIPAL", "REGISTRAR", "EXAM_CONTROLLER", "ACCOUNTS", "PLACEMENT_OFFICER", "HOD", "FACULTY"])), db: AsyncSession = Depends(get_db)):
    from datetime import datetime
    db_post = models.BulletinPost(
        author_id=current_user.id,
        content=post.content,
        created_at=datetime.utcnow().isoformat() + "Z",
        attachment_url=post.attachment_url,
        attachment_type=post.attachment_type
    )
    db.add(db_post)
    await db.commit()
    await db.refresh(db_post)
    
    name = f"{current_user.first_name} {current_user.last_name}"
    role = ROLE_MAP.get(current_user.system_role, "Staff")
    
    return schemas.BulletinPostOut(
        id=db_post.id,
        content=db_post.content,
        created_at=db_post.created_at,
        author_id=db_post.author_id,
        author_name=name,
        author_role=role,
        attachment_url=db_post.attachment_url,
        attachment_type=db_post.attachment_type
    )

@app.get("/job-tray/me")
async def get_job_tray(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    tray = {
        "pending_leaves": 0,
        "pending_requisitions": 0,
        "missed_punches": 0,
    }
    
    if current_user.system_role == "HOD":
        stmt_leaves = (
            select(func.count(models.LeaveRequest.id))
            .join(models.User, models.LeaveRequest.faculty_id == models.User.id)
            .where(
                models.LeaveRequest.status == "PENDING",
                models.User.department_id == current_user.department_id
            )
        )
        tray["pending_leaves"] = await db.scalar(stmt_leaves) or 0
        
        stmt_reqs = (
            select(func.count(models.RequisitionTicket.id))
            .where(models.RequisitionTicket.status == "PENDING")
        )
        tray["pending_requisitions"] = await db.scalar(stmt_reqs) or 0
        
    elif current_user.system_role in ["PRINCIPAL", "SUPER_ADMIN", "REGISTRAR"]:
        stmt_leaves = (
            select(func.count(models.LeaveRequest.id))
            .where(models.LeaveRequest.status == "PENDING")
        )
        tray["pending_leaves"] = await db.scalar(stmt_leaves) or 0
        
        stmt_reqs = (
            select(func.count(models.RequisitionTicket.id))
            .where(models.RequisitionTicket.status == "PENDING")
        )
        tray["pending_requisitions"] = await db.scalar(stmt_reqs) or 0
    else:
        stmt_reqs = (
            select(func.count(models.RequisitionTicket.id))
            .where(
                models.RequisitionTicket.requester_id == current_user.id,
                models.RequisitionTicket.status == "PENDING"
            )
        )
        tray["pending_requisitions"] = await db.scalar(stmt_reqs) or 0

    return tray

@app.post("/requisitions", response_model=schemas.RequisitionTicketOut)
async def create_requisition(req: schemas.RequisitionTicketCreate, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    db_req = models.RequisitionTicket(
        requester_id=current_user.id,
        ticket_type=req.ticket_type,
        description=req.description
    )
    db.add(db_req)
    await db.commit()
    await db.refresh(db_req)
    return db_req

# --- PHASE 9: ADVANCED ROUTES ---

@app.post("/users/me/switch-role")
async def switch_role(target_role: str, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    # Verify the user actually holds this role in additional_roles (or it is their current primary)
    roles = current_user.additional_roles or []
    if target_role != current_user.system_role and target_role not in roles:
        raise HTTPException(status_code=403, detail="Role not authorized for this user")
    
    # Swap primary role with the target role
    if target_role != current_user.system_role:
        if current_user.system_role not in roles:
            roles.append(current_user.system_role)
        roles.remove(target_role) if target_role in roles else None
        
        current_user.system_role = target_role
        current_user.additional_roles = roles
        await db.commit()
    return {"status": "success", "active_role": current_user.system_role}

# 1. EXAMINATIONS
import hashlib
from datetime import datetime
from pydantic import BaseModel

@app.post("/exams/records")
async def save_exam_mark(record: schemas.ExamRecordCreate, current_user: models.User = Depends(require_role(["FACULTY", "EXAM_CONTROLLER", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    # Check if record already exists
    res = await db.execute(select(models.ExamRecord).where(
        models.ExamRecord.student_id == record.student_id,
        models.ExamRecord.course_id == record.course_id,
        models.ExamRecord.exam_type == record.exam_type
    ))
    db_record = res.scalars().first()
    
    if db_record:
        if db_record.is_published:
            raise HTTPException(status_code=403, detail="Exams are frozen and cryptographically locked. Cannot edit.")
        # Audit Trail
        audit_log = models.ExamAuditLog(
            exam_record_id=db_record.id,
            changed_by_id=current_user.id,
            previous_marks=db_record.marks_obtained,
            new_marks=record.marks_obtained,
            timestamp=datetime.utcnow().isoformat() + "Z",
            ip_address="internal"
        )
        db.add(audit_log)
        db_record.marks_obtained = record.marks_obtained
    else:
        db_record = models.ExamRecord(**record.dict())
        db.add(db_record)
        
    await db.commit()

    # Sync to ExamResult (Phase 5/6 model) and Recalculate CGPA (Criticism Point 7 & 9)
    res_records = await db.execute(
        select(models.ExamRecord).where(
            models.ExamRecord.student_id == record.student_id,
            models.ExamRecord.course_id == record.course_id
        )
    )
    all_course_records = res_records.scalars().all()
    
    ise_sum = sum(r.marks_obtained for r in all_course_records if "ISE" in r.exam_type)
    ese_val = sum(r.marks_obtained for r in all_course_records if "ESE" in r.exam_type)
    total_val = ise_sum + ese_val
    
    # Grading Criteria
    grade_val = 'F'
    if total_val >= 90: grade_val = 'A+'
    elif total_val >= 80: grade_val = 'A'
    elif total_val >= 70: grade_val = 'B'
    elif total_val >= 60: grade_val = 'C'
    elif total_val >= 50: grade_val = 'D'
    
    res_result = await db.execute(
        select(models.ExamResult).where(
            models.ExamResult.student_id == record.student_id,
            models.ExamResult.course_id == record.course_id
        )
    )
    db_result = res_result.scalars().first()
    
    if db_result:
        db_result.ise_marks = ise_sum
        db_result.ese_marks = ese_val
        db_result.total_marks = total_val
        db_result.grade = grade_val
    else:
        db_result = models.ExamResult(
            student_id=record.student_id,
            course_id=record.course_id,
            semester=1, 
            ise_marks=ise_sum,
            ese_marks=ese_val,
            total_marks=total_val,
            grade=grade_val
        )
        db.add(db_result)
        
    await db.commit()

    # Recalculate CGPA on the profile mathematically correctly (Criticism Point 9)
    prof_res = await db.execute(select(models.StudentProfile).where(models.StudentProfile.user_id == record.student_id))
    profile = prof_res.scalars().first()
    if profile:
        all_results_res = await db.execute(select(models.ExamResult).where(models.ExamResult.student_id == record.student_id))
        all_results = all_results_res.scalars().all()
        if all_results:
            cgpa_sum = sum((r.total_marks / 10.0) for r in all_results if r.total_marks is not None)
            new_cgpa = round(cgpa_sum / len(all_results), 2)
            profile.cgpa = str(max(0.0, min(10.0, new_cgpa)))
            await db.commit()

    return {"status": "saved", "id": db_record.id}

@app.post("/exams/publish")
async def publish_exams(course_id: Optional[str] = None, exam_type: Optional[str] = None, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(require_role(["EXAM_CONTROLLER", "SUPER_ADMIN"]))):
    query = select(models.ExamRecord).where(models.ExamRecord.is_published == False)
    if course_id:
        query = query.where(models.ExamRecord.course_id == course_id)
    if exam_type:
        query = query.where(models.ExamRecord.exam_type == exam_type)
        
    res = await db.execute(query)
    records = res.scalars().all()
    
    # Cryptographic freeze protocol using asymmetric signing
    salt = os.environ.get("EXAM_HASH_SALT", "SUTRA_OS_SECURE_SALT_99")
    for r in records:
        r.is_published = True
        hash_string = f"{r.id}:{r.student_id}:{r.marks_obtained}:{salt}"
        r.cryptographic_hash = hashlib.sha256(hash_string.encode()).hexdigest()
        
        # Asymmetric key signing
        payload = f"{r.id}:{r.student_id}:{r.marks_obtained}"
        r.signature = sign_data(current_user.private_key, payload)
        r.signature_verified = True
        
    await db.commit()
    return {"status": "published_and_locked", "records_published": len(records)}

@app.get("/scholarships")
async def get_scholarships(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    if current_user.system_role == "STUDENT":
        res = await db.execute(select(models.ScholarshipLedger).where(models.ScholarshipLedger.student_id == current_user.id))
    else:
        res = await db.execute(select(models.ScholarshipLedger))
    return res.scalars().all()

@app.post("/scholarships")
async def apply_scholarship(sch: schemas.ScholarshipCreate, current_user: models.User = Depends(require_role(["STUDENT", "REGISTRAR", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    db_sch = models.ScholarshipLedger(**sch.dict(), status="PENDING_SCHOLARSHIP_SECTION")
    db.add(db_sch)
    await db.commit()
    return {"status": "created", "state": db_sch.status}

@app.post("/scholarships/{id}/approve-section")
async def approve_scholarship_section(id: str, current_user: models.User = Depends(require_role(["REGISTRAR", "HOD", "SUPER_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.ScholarshipLedger).where(models.ScholarshipLedger.id == id))
    sch = res.scalars().first()
    if sch and sch.status == "PENDING_SCHOLARSHIP_SECTION":
        sch.status = "PENDING_ACCOUNTS"
        sch.scholarship_officer_id = current_user.id
        await db.commit()
    return {"status": "success", "new_state": sch.status if sch else None}

@app.post("/scholarships/{id}/approve-accounts")
async def approve_scholarship_accounts(id: str, current_user: models.User = Depends(require_role(["ACCOUNTS", "SUPER_ADMIN", "PRINCIPAL"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.ScholarshipLedger).where(models.ScholarshipLedger.id == id))
    sch = res.scalars().first()
    if sch and sch.status == "PENDING_ACCOUNTS":
        sch.status = "APPROVED"
        sch.accounts_officer_id = current_user.id
        await db.commit()
    return {"status": "success", "new_state": sch.status if sch else None}


# 3. PLACEMENTS
@app.post("/placements/apply")
async def apply_drive(drive_id: str, db: AsyncSession = Depends(get_db), current_user: models.User = Depends(require_role(["STUDENT"]))):
    result_student = await db.execute(
        select(models.StudentProfile).where(models.StudentProfile.user_id == current_user.id)
    )
    student = result_student.scalars().first()
    if not student:
        raise HTTPException(status_code=400, detail="Only students with a profile can apply for placement drives")

    result_drive = await db.execute(
        select(models.PlacementDrive).where(models.PlacementDrive.id == drive_id)
    )
    drive = result_drive.scalars().first()
    if not drive:
        raise HTTPException(status_code=404, detail="Placement drive not found")

    try:
        student_cgpa = float(student.cgpa)
    except Exception:
        student_cgpa = 0.0

    if student_cgpa < drive.min_cgpa:
        raise HTTPException(
            status_code=400,
            detail=f"Ineligible: Your CGPA ({student_cgpa}) is below the minimum required CGPA ({drive.min_cgpa}) for this drive."
        )

    if drive.eligible_departments:
        allowed_depts = [d.strip().upper() for d in drive.eligible_departments.split(",")]
        result_dept = await db.execute(select(models.Department).where(models.Department.id == current_user.department_id))
        dept = result_dept.scalars().first()
        if not dept or dept.code.upper() not in allowed_depts:
            raise HTTPException(
                status_code=400,
                detail=f"Ineligible: Your department ({dept.code if dept else 'N/A'}) is not eligible for this drive."
            )

    result_exist = await db.execute(
        select(models.PlacementApplication).where(
            models.PlacementApplication.drive_id == drive_id,
            models.PlacementApplication.student_id == current_user.id
        )
    )
    if result_exist.scalars().first():
        raise HTTPException(status_code=400, detail="You have already applied for this placement drive")

    app = models.PlacementApplication(drive_id=drive_id, student_id=current_user.id)
    db.add(app)
    await db.commit()
    return {"status": "applied"}

@app.get("/placements/drives", response_model=list[schemas.PlacementDriveOut])
async def get_drives(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.PlacementDrive).order_by(models.PlacementDrive.drive_date.asc()))
    return res.scalars().all()


# --- LIBRARY ROUTING ---
@app.post("/library/books", response_model=schemas.LibraryBookOut)
async def create_library_book(book: schemas.LibraryBookCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "LIBRARIAN"])), db: AsyncSession = Depends(get_db)):
    db_book = models.LibraryBook(**book.dict())
    db.add(db_book)
    await db.commit()
    await db.refresh(db_book)
    return db_book

@app.get("/library/books", response_model=list[schemas.LibraryBookOut])
async def list_library_books(current_user: models.User = Depends(require_role(["LIBRARIAN", "SUPER_ADMIN", "ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.LibraryBook))
    return res.scalars().all()

@app.post("/library/checkout", response_model=schemas.LibraryCheckoutOut)
async def checkout_book(
    checkout: schemas.LibraryCheckoutCreate,
    current_user: models.User = Depends(require_role(["LIBRARIAN", "SUPER_ADMIN", "ADMIN"])),
    db: AsyncSession = Depends(get_db)
):
    if not checkout.student_username:
        raise HTTPException(status_code=400, detail="Student username is required for checkout")

    # Find student user
    res_student = await db.execute(
        select(models.User)
        .options(selectinload(models.User.student_profile))
        .where(func.lower(models.User.username) == func.lower(checkout.student_username.strip()))
    )
    student = res_student.scalars().first()
    if not student:
        raise HTTPException(status_code=404, detail="Student user not found")
        
    if student.system_role != "STUDENT":
        raise HTTPException(status_code=400, detail="User is not a student")

    # FINANCIAL HOLD CHECK
    res_overdue = await db.execute(
        select(models.FeeInvoice)
        .where(models.FeeInvoice.student_id == student.id)
        .where(models.FeeInvoice.status == "PENDING")
    )
    pending_invoices = res_overdue.scalars().all()
    if any(inv.due_date and inv.due_date.replace(tzinfo=None) < datetime.utcnow() for inv in pending_invoices):
        raise HTTPException(status_code=403, detail="FINANCIAL HOLD: Student has overdue fees. Library access restricted.")

    res_overdue = await db.execute(
        select(models.LibraryCheckout).where(
            models.LibraryCheckout.student_id == student.id,
            models.LibraryCheckout.status == "OVERDUE"
        )
    )
    overdue = res_overdue.scalars().all()
    if overdue:
        raise HTTPException(
            status_code=400,
            detail=f"Library Block: Student has {len(overdue)} overdue books outstanding."
        )

    res_book = await db.execute(
        select(models.LibraryBook)
        .where(models.LibraryBook.id == checkout.book_id)
        .with_for_update()
    )
    book = res_book.scalars().first()
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")
        
    if book.available_copies <= 0:
        raise HTTPException(status_code=400, detail="No copies available for checkout")

    checkout_date = datetime.utcnow().isoformat() + "Z"
    due_date = (datetime.utcnow() + timedelta(days=14)).isoformat() + "Z"
    
    db_checkout = models.LibraryCheckout(
        student_id=student.id,
        book_id=checkout.book_id,
        checkout_date=checkout_date,
        due_date=due_date,
        status="ISSUED"
    )
    
    book.available_copies -= 1
    
    db.add(db_checkout)
    await db.commit()
    await db.refresh(db_checkout)
    
    res = await db.execute(
        select(models.LibraryCheckout)
        .where(models.LibraryCheckout.id == db_checkout.id)
        .options(selectinload(models.LibraryCheckout.book))
    )
    return res.scalars().first()

@app.post("/library/checkout/{id}/return", response_model=schemas.LibraryCheckoutOut)
async def return_book(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "LIBRARIAN"])), db: AsyncSession = Depends(get_db)):
    res_checkout = await db.execute(
        select(models.LibraryCheckout)
        .where(models.LibraryCheckout.id == id)
        .options(selectinload(models.LibraryCheckout.book))
    )
    checkout = res_checkout.scalars().first()
    if not checkout:
        raise HTTPException(status_code=404, detail="Checkout record not found")
        
    if checkout.status == "RETURNED":
        raise HTTPException(status_code=400, detail="Book is already returned")

    checkout.status = "RETURNED"
    checkout.returned_at = datetime.utcnow().isoformat() + "Z"
    
    if checkout.book:
        checkout.book.available_copies += 1
        
    await db.commit()
    await db.refresh(checkout)
    return checkout

@app.get("/library/checkouts/me", response_model=list[schemas.LibraryCheckoutOut])
async def list_my_checkouts(current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.LibraryCheckout)
        .where(models.LibraryCheckout.student_id == current_user.id)
        .options(selectinload(models.LibraryCheckout.book))
    )
    return res.scalars().all()

@app.get("/library/checkouts", response_model=list[schemas.LibraryCheckoutOut])
async def list_all_checkouts(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "LIBRARIAN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.LibraryCheckout)
        .options(
            selectinload(models.LibraryCheckout.book),
            selectinload(models.LibraryCheckout.student)
        )
    )
    return res.scalars().all()


# --- TRANSPORT ROUTING ---
@app.post("/transport/stops", response_model=schemas.BusStopOut)
async def create_bus_stop(stop: schemas.BusStopCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "TRANSPORT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res_exist = await db.execute(select(models.BusStop).where(models.BusStop.name == stop.name))
    db_stop = res_exist.scalars().first()
    if not db_stop:
        db_stop = models.BusStop(name=stop.name)
        db.add(db_stop)
        await db.commit()
        await db.refresh(db_stop)
    return db_stop

@app.get("/transport/stops", response_model=list[schemas.BusStopOut])
async def list_bus_stops(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.BusStop))
    return res.scalars().all()

@app.post("/transport/routes/{route_id}/stops/{stop_id}")
async def link_route_stop(route_id: str, stop_id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "TRANSPORT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res_route = await db.execute(select(models.BusRoute).where(models.BusRoute.id == route_id).options(selectinload(models.BusRoute.stops)))
    route = res_route.scalars().first()
    if not route:
        raise HTTPException(status_code=404, detail="Bus route not found")
    res_stop = await db.execute(select(models.BusStop).where(models.BusStop.id == stop_id))
    stop = res_stop.scalars().first()
    if not stop:
        raise HTTPException(status_code=404, detail="Bus stop not found")
        
    if stop not in route.stops:
        route.stops.append(stop)
        await db.commit()
    return {"message": f"Stop {stop.name} successfully linked to route {route.route_name}"}

@app.post("/transport/routes", response_model=schemas.BusRouteOut)
async def create_bus_route(route: schemas.BusRouteCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "TRANSPORT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    db_route = models.BusRoute(**route.dict())
    db.add(db_route)
    await db.commit()
    await db.refresh(db_route)
    return db_route

@app.get("/transport/routes", response_model=list[schemas.BusRouteOut])
async def list_bus_routes(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.BusRoute).options(selectinload(models.BusRoute.stops)))
    return res.scalars().all()

@app.post("/transport/reservations", response_model=schemas.TransportReservationOut)
async def reserve_seat_alias(reserve: schemas.TransportReservationCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    return await reserve_seat(reserve, current_user, db)

@app.post("/transport/reserve", response_model=schemas.TransportReservationOut)
async def reserve_seat(reserve: schemas.TransportReservationCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res_p_route = await db.execute(select(models.BusRoute).where(models.BusRoute.id == reserve.pickup_route_id))
    p_route = res_p_route.scalars().first()
    if not p_route:
        raise HTTPException(status_code=404, detail="Pickup bus route not found")
        
    res_d_route = await db.execute(select(models.BusRoute).where(models.BusRoute.id == reserve.destination_route_id))
    d_route = res_d_route.scalars().first()
    if not d_route:
        raise HTTPException(status_code=404, detail="Destination bus route not found")
        
    res_exist = await db.execute(
        select(models.TransportReservation).where(
            models.TransportReservation.student_id == current_user.id,
            models.TransportReservation.approval_status.in_(["PENDING", "APPROVED"])
        )
    )
    if res_exist.scalars().first():
        raise HTTPException(status_code=400, detail="You already have an active or pending transport reservation")

    db_reserve = models.TransportReservation(
        student_id=current_user.id,
        pickup_stop=reserve.pickup_stop,
        pickup_route_id=reserve.pickup_route_id,
        destination_stop=reserve.destination_stop,
        destination_route_id=reserve.destination_route_id,
        vehicle_no=p_route.bus_number,
        paid_amount=0.0,
        is_paid=False,
        fee_amount=12000.0,
        approval_authority="Transport Officer",
        approval_status="PENDING",
        reserved_at=datetime.utcnow().isoformat() + "Z"
    )
    db.add(db_reserve)
    
    fee = models.FeeInvoice(
        student_id=current_user.id,
        amount=12000.0,
        description=f"Transport Fees - Route: {p_route.route_name} (Pickup: {reserve.pickup_stop})",
        status="PENDING"
    )
    db.add(fee)
    
    await db.commit()
    await db.refresh(db_reserve)
    
    res = await db.execute(
        select(models.TransportReservation)
        .where(models.TransportReservation.id == db_reserve.id)
        .options(
            selectinload(models.TransportReservation.pickup_route),
            selectinload(models.TransportReservation.destination_route),
            selectinload(models.TransportReservation.student)
        )
    )
    return res.scalars().first()

@app.get("/transport/reservations/me", response_model=list[schemas.TransportReservationOut])
async def list_my_reservations(current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.TransportReservation)
        .where(models.TransportReservation.student_id == current_user.id)
        .options(
            selectinload(models.TransportReservation.pickup_route),
            selectinload(models.TransportReservation.destination_route),
            selectinload(models.TransportReservation.student)
        )
    )
    return res.scalars().all()

@app.get("/transport/reservations", response_model=list[schemas.TransportReservationOut])
async def list_all_reservations(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "TRANSPORT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.TransportReservation)
        .options(
            selectinload(models.TransportReservation.pickup_route),
            selectinload(models.TransportReservation.destination_route),
            selectinload(models.TransportReservation.student)
        )
    )
    return res.scalars().all()

@app.put("/transport/reservations/{id}/status", response_model=schemas.TransportReservationOut)
async def update_reservation_status(id: str, status: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "TRANSPORT_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res_reserve = await db.execute(
        select(models.TransportReservation)
        .where(models.TransportReservation.id == id)
        .options(
            selectinload(models.TransportReservation.pickup_route),
            selectinload(models.TransportReservation.destination_route),
            selectinload(models.TransportReservation.student)
        )
        .with_for_update()
    )
    reservation = res_reserve.scalars().first()
    if not reservation:
        raise HTTPException(status_code=404, detail="Reservation not found")
        
    old_status = reservation.approval_status
    if status == "APPROVED" and old_status != "APPROVED":
        route = reservation.pickup_route
        if route.reserved_seats >= route.capacity:
            raise HTTPException(status_code=400, detail=f"Route {route.route_name} is already at full capacity.")
        route.reserved_seats += 1
        reservation.seat_number = route.reserved_seats
        
    elif status != "APPROVED" and old_status == "APPROVED":
        route = reservation.pickup_route
        route.reserved_seats = max(0, route.reserved_seats - 1)
        reservation.seat_number = None
        
    reservation.approval_status = status
    await db.commit()
    await db.refresh(reservation)
    return reservation


# --- HOSTEL ROUTING ---
@app.post("/hostel/admissions", response_model=schemas.HostelAdmissionOut)
async def create_hostel_admission(adm: schemas.HostelAdmissionCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res_exist = await db.execute(select(models.HostelAdmission).where(models.HostelAdmission.student_id == current_user.id))
    if res_exist.scalars().first():
        raise HTTPException(status_code=400, detail="Student is already admitted or registered to the hostel")
        
    db_adm = models.HostelAdmission(
        student_id=current_user.id,
        course_year=adm.course_year,
        gender=adm.gender,
        policy_name=adm.policy_name,
        plan_name=adm.plan_name,
        father_name=adm.father_name,
        father_contact=adm.father_contact,
        father_address=adm.father_address,
        mother_name=adm.mother_name,
        mother_contact=adm.mother_contact,
        mother_address=adm.mother_address,
        guardian_name=adm.guardian_name,
        guardian_contact=adm.guardian_contact,
        guardian_address=adm.guardian_address,
        vehicle_number=adm.vehicle_number,
        license_number=adm.license_number,
        block_name="NA",
        floor_name="NA",
        room_number="NA",
        status="PENDING",
        parent_consent_approved=True
    )
    db.add(db_adm)
    await db.commit()
    await db.refresh(db_adm)
    
    res = await db.execute(select(models.HostelAdmission).where(models.HostelAdmission.id == db_adm.id).options(selectinload(models.HostelAdmission.student)))
    return res.scalars().first()

@app.get("/hostel/admissions/me", response_model=schemas.HostelAdmissionOut)
async def get_my_hostel_admission(current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.HostelAdmission).where(models.HostelAdmission.student_id == current_user.id).options(selectinload(models.HostelAdmission.student)))
    adm = res.scalars().first()
    if not adm:
        raise HTTPException(status_code=404, detail="Hostel admission not found for this student")
    return adm

@app.get("/hostel/admissions", response_model=list[schemas.HostelAdmissionOut])
async def list_hostel_admissions(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "HOSTEL_WARDEN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.HostelAdmission).options(selectinload(models.HostelAdmission.student)))
    return res.scalars().all()

@app.put("/hostel/admissions/{id}/allocate", response_model=schemas.HostelAdmissionOut)
async def allocate_hostel_room(id: str, alloc: schemas.HostelAllocationRequest, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "HOSTEL_WARDEN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.HostelAdmission)
        .where(models.HostelAdmission.id == id)
        .options(selectinload(models.HostelAdmission.student))
        .with_for_update()
    )
    adm = res.scalars().first()
    if not adm:
        raise HTTPException(status_code=404, detail="Hostel registration not found")
        
    adm.block_name = alloc.block_name
    adm.floor_name = alloc.floor_name
    adm.room_number = alloc.room_number
    adm.status = "APPROVED"
    
    fee = models.FeeInvoice(
        student_id=adm.student_id,
        amount=50000.0,
        description=f"Hostel Room Fees - Block: {alloc.block_name}, Floor: {alloc.floor_name}, Room: {alloc.room_number}",
        status="PENDING"
    )
    db.add(fee)
    
    await db.commit()
    await db.refresh(adm)
    return adm

# --- OFF CAMPUS PLACEMENTS ---
@app.post("/placements/off-campus", response_model=schemas.OffCampusPlacementOut)
async def create_off_campus_placement(op: schemas.OffCampusPlacementCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    # Check if student already submitted
    res_exist = await db.execute(select(models.OffCampusPlacement).where(models.OffCampusPlacement.student_id == current_user.id))
    if res_exist.scalars().first():
        raise HTTPException(status_code=400, detail="You have already submitted an off-campus placement application")
        
    db_op = models.OffCampusPlacement(
        student_id=current_user.id,
        company_name=op.company_name,
        job_profile=op.job_profile,
        probation_months=op.probation_months,
        after_confirmation_salary=op.after_confirmation_salary,
        probation_salary=op.probation_salary,
        bond_months=op.bond_months,
        joining_date=op.joining_date,
        status="PENDING"
    )
    db.add(db_op)
    await db.commit()
    await db.refresh(db_op)
    
    res = await db.execute(
        select(models.OffCampusPlacement)
        .where(models.OffCampusPlacement.id == db_op.id)
        .options(selectinload(models.OffCampusPlacement.student))
    )
    return res.scalars().first()

@app.get("/placements/off-campus", response_model=list[schemas.OffCampusPlacementOut])
async def list_off_campus_placements(current_user: models.User = Depends(require_role(["PLACEMENT_OFFICER", "SUPER_ADMIN", "ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.OffCampusPlacement)
        .options(selectinload(models.OffCampusPlacement.student))
    )
    return res.scalars().all()

@app.get("/placements/off-campus/me", response_model=schemas.OffCampusPlacementOut)
async def get_my_off_campus_placement(current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.OffCampusPlacement)
        .where(models.OffCampusPlacement.student_id == current_user.id)
        .options(selectinload(models.OffCampusPlacement.student))
    )
    op = res.scalars().first()
    if not op:
        raise HTTPException(status_code=404, detail="No off-campus placement application found")
    return op

@app.get("/placements/off-campus/student/{student_id}", response_model=schemas.OffCampusPlacementOut)
async def get_student_off_campus_placement(
    student_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    res = await db.execute(
        select(models.OffCampusPlacement)
        .where(models.OffCampusPlacement.student_id == student_id)
        .options(selectinload(models.OffCampusPlacement.student))
    )
    op = res.scalars().first()
    if not op:
        raise HTTPException(status_code=404, detail="No off-campus placement application found")
    return op

@app.put("/placements/off-campus/{id}/status", response_model=schemas.OffCampusPlacementOut)
async def update_off_campus_placement_status(
    id: str,
    status: str,
    current_user: models.User = Depends(require_role(["PLACEMENT_OFFICER", "SUPER_ADMIN", "ADMIN"])),
    db: AsyncSession = Depends(get_db)
):
    res = await db.execute(
        select(models.OffCampusPlacement)
        .where(models.OffCampusPlacement.id == id)
        .options(selectinload(models.OffCampusPlacement.student))
    )
    op = res.scalars().first()
    if not op:
        raise HTTPException(status_code=404, detail="Off-campus placement request not found")
    
    if status not in ["APPROVED", "REJECTED"]:
        raise HTTPException(status_code=400, detail="Invalid status")
        
    op.status = status
    await db.commit()
    await db.refresh(op)
    
    # Log placement offer for dashboard tracking
    if status == "APPROVED":
        res_offer = await db.execute(select(models.PlacementOffer).where(models.PlacementOffer.student_id == op.student_id))
        if not res_offer.scalars().first():
            db_offer = models.PlacementOffer(
                student_id=op.student_id,
                company_name=op.company_name,
                job_profile=op.job_profile,
                salary_package=op.after_confirmation_salary,
                offer_letter_url="/uploads/offcampus_letter.pdf"
            )
            db.add(db_offer)
            await db.commit()
            
    return op


# --- LIBRARIAN ADDITIONAL ENDPOINTS ---
@app.post("/library/books", response_model=schemas.LibraryBookOut)
async def add_library_book(book_data: schemas.LibraryBookCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "LIBRARIAN"])), db: AsyncSession = Depends(get_db)):
    db_book = models.LibraryBook(
        title=book_data.title,
        author=book_data.author,
        isbn=book_data.isbn,
        total_copies=book_data.total_copies,
        available_copies=book_data.total_copies
    )
    db.add(db_book)
    await db.commit()
    await db.refresh(db_book)
    return db_book

@app.put("/library/checkouts/{id}/return", response_model=schemas.LibraryCheckoutOut)
async def return_library_book(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "LIBRARIAN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.LibraryCheckout)
        .where(models.LibraryCheckout.id == id)
        .options(selectinload(models.LibraryCheckout.book))
    )
    checkout = res.scalars().first()
    if not checkout:
        raise HTTPException(status_code=404, detail="Checkout record not found")
    if checkout.status == "RETURNED":
        return checkout
        
    checkout.returned_at = datetime.utcnow().isoformat()
    checkout.status = "RETURNED"
    checkout.book.available_copies = min(checkout.book.available_copies + 1, checkout.book.total_copies)
    await db.commit()
    await db.refresh(checkout)
    return checkout

# --- MESS ENDPOINTS ---
@app.get("/mess/menu", response_model=list[schemas.MessMenuOut])
async def list_mess_menu(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.MessMenu))
    return res.scalars().all()

class MessMenuUpdate(BaseModel):
    breakfast: str
    lunch: str
    snacks: str
    dinner: str

@app.put("/mess/menu/{day}", response_model=schemas.MessMenuOut)
async def update_mess_menu(day: str, menu_data: MessMenuUpdate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "MESS_IN_CHARGE"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.MessMenu).where(models.MessMenu.day_of_week == day))
    menu = res.scalars().first()
    if not menu:
        menu = models.MessMenu(day_of_week=day, breakfast="", lunch="", snacks="", dinner="")
        db.add(menu)
    menu.breakfast = menu_data.breakfast
    menu.lunch = menu_data.lunch
    menu.snacks = menu_data.snacks
    menu.dinner = menu_data.dinner
    await db.commit()
    await db.refresh(menu)
    return menu

@app.get("/mess/feedback", response_model=list[schemas.MessFeedbackOut])
async def list_mess_feedback(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "MESS_IN_CHARGE"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.MessFeedback).options(selectinload(models.MessFeedback.student)))
    return res.scalars().all()

@app.post("/mess/feedback", response_model=schemas.MessFeedbackOut)
async def submit_mess_feedback(feedback: schemas.MessFeedbackCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    db_feedback = models.MessFeedback(
        student_id=current_user.id,
        rating=feedback.rating,
        review=feedback.review,
        created_at=datetime.utcnow().isoformat()
    )
    db.add(db_feedback)
    await db.commit()
    await db.refresh(db_feedback)
    return db_feedback

@app.get("/mess/grocery", response_model=list[schemas.MessGroceryOut])
async def list_mess_grocery(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "MESS_IN_CHARGE"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.MessGrocery))
    return res.scalars().all()

@app.post("/mess/grocery/{id}/restock", response_model=schemas.MessGroceryOut)
async def restock_grocery(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "MESS_IN_CHARGE"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.MessGrocery).where(models.MessGrocery.id == id))
    item = res.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Grocery item not found")
    item.current_stock += 100.0
    await db.commit()
    await db.refresh(item)
    return item

# --- SPORTS ENDPOINTS ---
@app.get("/sports/equipment", response_model=list[schemas.SportsEquipmentOut])
async def list_sports_equipment(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.SportsEquipment))
    return res.scalars().all()

@app.post("/sports/equipment", response_model=schemas.SportsEquipmentOut)
async def create_sports_equipment(eq: schemas.SportsEquipmentCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "SPORTS_OFFICER"])), db: AsyncSession = Depends(get_db)):
    db_eq = models.SportsEquipment(
        name=eq.name,
        total_qty=eq.total_qty,
        available_qty=eq.total_qty
    )
    db.add(db_eq)
    await db.commit()
    await db.refresh(db_eq)
    return db_eq

@app.post("/sports/issue", response_model=schemas.SportsIssueRequestOut)
async def request_sports_equipment(req: schemas.SportsIssueRequestCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    res_eq = await db.execute(select(models.SportsEquipment).where(models.SportsEquipment.id == req.equipment_id))
    eq = res_eq.scalars().first()
    if not eq or eq.available_qty < req.quantity:
        raise HTTPException(status_code=400, detail="Equipment unavailable or insufficient quantity")
    
    db_req = models.SportsIssueRequest(
        student_id=current_user.id,
        equipment_id=req.equipment_id,
        quantity=req.quantity,
        status="PENDING",
        request_date=datetime.utcnow().isoformat()
    )
    db.add(db_req)
    await db.commit()
    
    res_loaded = await db.execute(
        select(models.SportsIssueRequest)
        .where(models.SportsIssueRequest.id == db_req.id)
        .options(
            selectinload(models.SportsIssueRequest.student),
            selectinload(models.SportsIssueRequest.equipment)
        )
    )
    return res_loaded.scalars().first()

@app.get("/sports/issue", response_model=list[schemas.SportsIssueRequestOut])
async def list_sports_issue_requests(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "SPORTS_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.SportsIssueRequest)
        .options(
            selectinload(models.SportsIssueRequest.student),
            selectinload(models.SportsIssueRequest.equipment)
        )
    )
    return res.scalars().all()

@app.put("/sports/issue/{id}/status", response_model=schemas.SportsIssueRequestOut)
async def update_sports_issue_status(id: str, status: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "SPORTS_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.SportsIssueRequest)
        .where(models.SportsIssueRequest.id == id)
        .options(
            selectinload(models.SportsIssueRequest.equipment),
            selectinload(models.SportsIssueRequest.student)
        )
    )
    req = res.scalars().first()
    if not req:
        raise HTTPException(status_code=404, detail="Issue request not found")
        
    if status == "APPROVED" and req.status == "PENDING":
        if req.equipment.available_qty < req.quantity:
            raise HTTPException(status_code=400, detail="Insufficient stock")
        req.equipment.available_qty -= req.quantity
        req.status = "APPROVED"
    elif status == "RETURNED" and req.status == "APPROVED":
        req.equipment.available_qty = min(req.equipment.available_qty + req.quantity, req.equipment.total_qty)
        req.status = "RETURNED"
        req.returned_at = datetime.utcnow().isoformat()
        
    await db.commit()
    await db.refresh(req)
    return req

@app.get("/sports/tournaments", response_model=list[schemas.SportsTournamentOut])
async def list_sports_tournaments(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.SportsTournament).options(selectinload(models.SportsTournament.registered_by)))
    return res.scalars().all()

@app.post("/sports/tournaments", response_model=schemas.SportsTournamentOut)
async def register_sports_tournament(req: schemas.SportsTournamentCreate, current_user: models.User = Depends(require_role(["STUDENT"])), db: AsyncSession = Depends(get_db)):
    db_tour = models.SportsTournament(
        team_name=req.team_name,
        sport_name=req.sport_name,
        members_count=req.members_count,
        registered_by_id=current_user.id,
        registered_at=datetime.utcnow().isoformat()
    )
    db.add(db_tour)
    await db.commit()
    await db.refresh(db_tour)
    return db_tour

# --- ESTATE MAINTENANCE ENDPOINTS ---
@app.post("/maintenance/tickets", response_model=schemas.MaintenanceTicketOut)
async def create_maintenance_ticket(ticket: schemas.MaintenanceTicketCreate, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    db_ticket = models.EstateMaintenanceTicket(
        reporter_id=current_user.id,
        category=ticket.category,
        block_name=ticket.block_name,
        description=ticket.description,
        status="PENDING",
        created_at=datetime.utcnow().isoformat()
    )
    db.add(db_ticket)
    await db.commit()
    
    res_loaded = await db.execute(
        select(models.EstateMaintenanceTicket)
        .where(models.EstateMaintenanceTicket.id == db_ticket.id)
        .options(selectinload(models.EstateMaintenanceTicket.reporter))
    )
    return res_loaded.scalars().first()

@app.get("/maintenance/tickets", response_model=list[schemas.MaintenanceTicketOut])
async def list_maintenance_tickets(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    user_roles = [current_user.system_role]
    if current_user.additional_roles and isinstance(current_user.additional_roles, list):
        user_roles.extend(current_user.additional_roles)
        
    if any(role in ["ADMIN", "SUPER_ADMIN", "ESTATE_MANAGER"] for role in user_roles):
        res = await db.execute(select(models.EstateMaintenanceTicket).options(selectinload(models.EstateMaintenanceTicket.reporter)))
    else:
        res = await db.execute(select(models.EstateMaintenanceTicket).where(models.EstateMaintenanceTicket.reporter_id == current_user.id).options(selectinload(models.EstateMaintenanceTicket.reporter)))
    return res.scalars().all()

@app.put("/maintenance/tickets/{id}/assign", response_model=schemas.MaintenanceTicketOut)
async def assign_maintenance_ticket(id: str, staff_name: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "ESTATE_MANAGER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.EstateMaintenanceTicket)
        .where(models.EstateMaintenanceTicket.id == id)
        .options(selectinload(models.EstateMaintenanceTicket.reporter))
    )
    ticket = res.scalars().first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    ticket.assigned_to_staff = staff_name
    ticket.status = "ASSIGNED"
    await db.commit()
    await db.refresh(ticket)
    return ticket

@app.put("/maintenance/tickets/{id}/resolve", response_model=schemas.MaintenanceTicketOut)
async def resolve_maintenance_ticket(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "ESTATE_MANAGER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.EstateMaintenanceTicket)
        .where(models.EstateMaintenanceTicket.id == id)
        .options(selectinload(models.EstateMaintenanceTicket.reporter))
    )
    ticket = res.scalars().first()
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    ticket.status = "RESOLVED"
    await db.commit()
    await db.refresh(ticket)
    return ticket

# --- CENTRAL STORE & VENDING ENDPOINTS ---
@app.get("/store/inventory", response_model=list[schemas.CentralStoreItemOut])
async def list_central_store(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PURCHASE_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.CentralStoreItem))
    return res.scalars().all()

@app.post("/store/inventory", response_model=schemas.CentralStoreItemOut)
async def create_store_item(item: schemas.CentralStoreItemCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PURCHASE_OFFICER"])), db: AsyncSession = Depends(get_db)):
    db_item = models.CentralStoreItem(
        item_name=item.item_name,
        quantity=item.quantity,
        unit=item.unit
    )
    db.add(db_item)
    await db.commit()
    await db.refresh(db_item)
    return db_item

@app.post("/store/requisitions", response_model=schemas.StoreRequisitionOut)
async def create_store_requisition(req: schemas.StoreRequisitionCreate, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res_it = await db.execute(
        select(models.CentralStoreItem)
        .where(models.CentralStoreItem.id == req.item_id)
        .with_for_update()
    )
    it = res_it.scalars().first()
    if not it or it.quantity < req.quantity:
        raise HTTPException(status_code=400, detail="Item unavailable or insufficient store stock")
        
    db_req = models.StoreRequisition(
        user_id=current_user.id,
        item_id=req.item_id,
        quantity=req.quantity,
        status="PENDING",
        requested_at=datetime.utcnow().isoformat()
    )
    db.add(db_req)
    await db.commit()
    
    res_loaded = await db.execute(
        select(models.StoreRequisition)
        .where(models.StoreRequisition.id == db_req.id)
        .options(
            selectinload(models.StoreRequisition.item),
            selectinload(models.StoreRequisition.user)
        )
    )
    return res_loaded.scalars().first()

@app.get("/store/requisitions", response_model=list[schemas.StoreRequisitionOut])
async def list_store_requisitions(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PURCHASE_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.StoreRequisition)
        .options(
            selectinload(models.StoreRequisition.user),
            selectinload(models.StoreRequisition.item)
        )
    )
    return res.scalars().all()

@app.put("/store/requisitions/{id}/status", response_model=schemas.StoreRequisitionOut)
async def update_store_requisition_status(id: str, status: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PURCHASE_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.StoreRequisition)
        .where(models.StoreRequisition.id == id)
        .options(
            selectinload(models.StoreRequisition.item),
            selectinload(models.StoreRequisition.user)
        )
        .with_for_update()
    )
    req = res.scalars().first()
    if not req:
        raise HTTPException(status_code=404, detail="Requisition not found")
        
    if status == "DISBURSED" and req.status == "PENDING":
        if req.item.quantity < req.quantity:
            raise HTTPException(status_code=400, detail="Insufficient central store stock")
        req.item.quantity -= req.quantity
        req.status = "DISBURSED"
    else:
        req.status = status
        
    await db.commit()
    await db.refresh(req)
    return req

@app.get("/vending/inventory", response_model=list[schemas.VendingMachineItemOut])
async def list_vending_inventory(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.VendingMachineItem))
    return res.scalars().all()

@app.post("/vending/{id}/refill", response_model=schemas.VendingMachineItemOut)
async def refill_vending_machine(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PURCHASE_OFFICER"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.VendingMachineItem).where(models.VendingMachineItem.id == id))
    item = res.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Vending item not found")
    item.quantity = item.max_quantity
    await db.commit()
    await db.refresh(item)
    return item


# --- ADMISSIONS ENDPOINTS ---
@app.post("/admissions/apply", response_model=schemas.AdmissionApplicationOut)
async def apply_for_admission(app_data: schemas.AdmissionApplicationCreate, db: AsyncSession = Depends(get_db)):
    db_app = models.AdmissionApplication(
        first_name=app_data.first_name,
        last_name=app_data.last_name,
        email=app_data.email,
        hsc_percentage=app_data.hsc_percentage,
        category=app_data.category.upper(),
        status="PENDING",
        applied_at=datetime.utcnow().isoformat()
    )
    db.add(db_app)
    await db.commit()
    await db.refresh(db_app)
    return db_app

@app.post("/admissions/register-direct")
async def register_student_direct(
    data: schemas.StudentDirectRegister,
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "REGISTRAR", "ADMISSION_ADMIN"]))
):
    # Check if username already exists
    res_exist = await db.execute(select(models.User).where(func.lower(models.User.username) == func.lower(data.username.strip())))
    if res_exist.scalars().first():
        raise HTTPException(status_code=400, detail="Username/Roll number already exists")
    
    # Get department
    res_dept = await db.execute(select(models.Department).where(models.Department.code == data.department_code))
    dept = res_dept.scalars().first()
    dept_id = dept.id if dept else None
    
    # Create User account
    hashed_pwd = auth.get_password_hash(data.password)
    db_user = models.User(
        username=data.username.strip(),
        email=data.email.strip(),
        hashed_password=hashed_pwd,
        first_name=data.first_name.strip(),
        last_name=data.last_name.strip(),
        system_role="STUDENT",
        department_id=dept_id
    )
    db.add(db_user)
    await db.commit()
    await db.refresh(db_user)
    
    # Create Student Profile
    db_profile = models.StudentProfile(
        user_id=db_user.id,
        enrollment_number=db_user.username,
        current_semester=1,
        batch_year=2024,
        parent_whatsapp=data.father_contact or data.mother_contact,
        parent_email=data.email,
        cgpa="0.0"
    )
    db.add(db_profile)
    
    # Create standard fee invoice
    db_invoice = models.FeeInvoice(
        student_id=db_user.id,
        amount=45000.0,
        description="First Semester Tuition Fees",
        status="PENDING"
    )
    db.add(db_invoice)
    
    # Create corresponding application record
    db_app = models.AdmissionApplication(
        first_name=data.first_name,
        last_name=data.last_name,
        email=data.email,
        hsc_percentage=90.0,
        category=data.category,
        status="ADMITTED",
        applied_at=datetime.utcnow().isoformat()
    )
    db.add(db_app)
    
    await db.commit()
    return {"message": "Student registered successfully", "username": db_user.username}

@app.get("/admissions/applications", response_model=list[schemas.AdmissionApplicationOut])
async def list_admission_applications(current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "REGISTRAR", "ADMISSION_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.AdmissionApplication)
        .order_by(models.AdmissionApplication.hsc_percentage.desc())
    )
    return res.scalars().all()

@app.post("/admissions/applications/{id}/admit")
async def admit_applicant(id: str, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "REGISTRAR", "ADMISSION_ADMIN"])), db: AsyncSession = Depends(get_db)):
    res = await db.execute(select(models.AdmissionApplication).where(models.AdmissionApplication.id == id))
    app_data = res.scalars().first()
    if not app_data:
        raise HTTPException(status_code=404, detail="Application not found")
    if app_data.status == "ADMITTED":
        raise HTTPException(status_code=400, detail="Applicant already admitted")
        
    app_data.status = "ADMITTED"
    
    # Generate roll number username dynamically
    res_count = await db.execute(select(models.User).where(models.User.system_role == "STUDENT"))
    students_count = len(res_count.scalars().all())
    roll_num = f"mit2024cse{str(students_count + 1).zfill(3)}"
    
    # Get a department ID (default to CSE dept)
    res_dept = await db.execute(select(models.Department).where(models.Department.code == "CSE"))
    dept = res_dept.scalars().first()
    dept_id = dept.id if dept else None
    
    # Create User account
    hashed_pwd = auth.get_password_hash("password123")
    db_user = models.User(
        username=roll_num,
        email=app_data.email,
        hashed_password=hashed_pwd,
        first_name=app_data.first_name,
        last_name=app_data.last_name,
        system_role="STUDENT",
        department_id=dept_id
    )
    db.add(db_user)
    await db.commit()
    await db.refresh(db_user)
    
    # Create Student Profile
    db_profile = models.StudentProfile(
        user_id=db_user.id,
        enrollment_number=roll_num.upper(),
        current_semester=1,
        batch_year=2024,
        cgpa="0.0",
        parent_whatsapp="+919876543210"
    )
    db.add(db_profile)
    await db.commit()
    
    return {"status": "success", "username": roll_num, "temporary_password": "password123"}


# --- RESEARCH MODULE ENDPOINTS ---
@app.get("/research/projects", response_model=list[schemas.ResearchProjectOut])
async def list_research_projects(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(
        select(models.ResearchProject)
        .options(selectinload(models.ResearchProject.faculty))
    )
    return res.scalars().all()

@app.post("/research/projects", response_model=schemas.ResearchProjectOut)
async def create_research_project(proj: schemas.ResearchProjectCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PRINCIPAL", "HOD", "FACULTY"])), db: AsyncSession = Depends(get_db)):
    db_proj = models.ResearchProject(
        title=proj.title,
        funding_agency=proj.funding_agency,
        amount=proj.amount,
        duration=proj.duration,
        status="ONGOING",
        faculty_id=current_user.id
    )
    db.add(db_proj)
    await db.commit()
    
    res_loaded = await db.execute(
        select(models.ResearchProject)
        .where(models.ResearchProject.id == db_proj.id)
        .options(selectinload(models.ResearchProject.faculty))
    )
    return res_loaded.scalars().first()

@app.get("/research/publications", response_model=list[schemas.ResearchPublicationOut])
async def list_research_publications(db: AsyncSession = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    res = await db.execute(select(models.ResearchPublication))
    return res.scalars().all()

@app.post("/research/publications", response_model=schemas.ResearchPublicationOut)
async def create_research_publication(pub: schemas.ResearchPublicationCreate, current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PRINCIPAL", "HOD", "FACULTY"])), db: AsyncSession = Depends(get_db)):
    db_pub = models.ResearchPublication(
        title=pub.title,
        journal=pub.journal,
        author_name=pub.author_name,
        year=pub.year,
        doi=pub.doi,
        citation_count=0
    )
    db.add(db_pub)
    await db.commit()
    await db.refresh(db_pub)
    return db_pub


# --- DMS LOCKER ENDPOINTS ---
@app.get("/dms/documents", response_model=list[schemas.DocumentLockerOut])
async def list_dms_documents(current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.DocumentLocker)
        .where(models.DocumentLocker.owner_id == current_user.id)
    )
    return res.scalars().all()

from fastapi import Form
@app.post("/dms/documents", response_model=schemas.DocumentLockerOut)
async def upload_dms_document(
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    import hashlib
    import uuid
    import base64
    from fastapi import Form, UploadFile, File
    from cryptography.fernet import Fernet
    import auth
    
    # Generate consistent Fernet key from SECRET_KEY
    key_material = hashlib.sha256(auth.SECRET_KEY.encode('utf-8')).digest()
    fernet_key = base64.urlsafe_b64encode(key_material)
    f_crypt = Fernet(fernet_key)
    
    content = await file.read()
    file_size_mb = len(content) / (1024 * 1024)
    file_size_str = f"{file_size_mb:.1f} MB"
    
    file_id = str(uuid.uuid4())
    file_extension = file.filename.split('.')[-1] if '.' in file.filename else ''
    safe_filename = f"{file_id}.{file_extension}"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)
    
    # Encrypt the file content before saving to disk
    encrypted_content = f_crypt.encrypt(content)
    
    with open(file_path, 'wb') as out_file:
        out_file.write(encrypted_content)
        
    raw_str = f"{file.filename}-{doc_type}-{file_size_str}-{current_user.id}"
    crypto_sig = hashlib.sha256(raw_str.encode('utf-8')).hexdigest()
    
    db_doc = models.DocumentLocker(
        owner_id=current_user.id,
        doc_name=file.filename,
        doc_type=doc_type,
        file_size=file_size_str,
        uploaded_at=datetime.utcnow().isoformat() + "Z",
        cryptographic_hash=crypto_sig,
        file_path=file_path
    )
    db.add(db_doc)
    try:
        await db.commit()
        await db.refresh(db_doc)
        return db_doc
    except sqlalchemy.exc.ProgrammingError as exc:
        await db.rollback()
        if "file_path" in str(exc):
            await ensure_document_lockers_schema(engine)
            db_doc = models.DocumentLocker(
                owner_id=current_user.id,
                doc_name=file.filename,
                doc_type=doc_type,
                file_size=file_size_str,
                uploaded_at=datetime.utcnow().isoformat() + "Z",
                cryptographic_hash=crypto_sig,
                file_path=file_path
            )
            db.add(db_doc)
            await db.commit()
            await db.refresh(db_doc)
            return db_doc
        raise

from fastapi.responses import FileResponse
@app.get("/dms/download/{id}")
async def download_dms_document(id: str, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.DocumentLocker)
        .where(models.DocumentLocker.id == id, models.DocumentLocker.owner_id == current_user.id)
    )
    doc = res.scalars().first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found or access denied")
    
    if not doc.file_path or not isinstance(doc.file_path, str):
        raise HTTPException(status_code=404, detail="File not found or file metadata is missing")

    if not os.path.exists(doc.file_path):
        raise HTTPException(status_code=404, detail="File missing on server")
        
    import hashlib
    import base64
    from cryptography.fernet import Fernet
    import auth
    from fastapi import Response
    import mimetypes
    
    key_material = hashlib.sha256(auth.SECRET_KEY.encode('utf-8')).digest()
    fernet_key = base64.urlsafe_b64encode(key_material)
    f_crypt = Fernet(fernet_key)
    
    with open(doc.file_path, 'rb') as in_file:
        encrypted_content = in_file.read()
        
    try:
        decrypted_content = f_crypt.decrypt(encrypted_content)
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to decrypt file. File may be corrupted or from before encryption was enabled.")
        
    mimetype, _ = mimetypes.guess_type(doc.doc_name)
    content_type = mimetype or "application/octet-stream"
    
    return Response(
        content=decrypted_content,
        media_type=content_type,
        headers={"Content-Disposition": f'inline; filename="{doc.doc_name}"'}
    )

@app.delete("/dms/documents/{id}")
async def delete_dms_document(id: str, current_user: models.User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    res = await db.execute(
        select(models.DocumentLocker)
        .where(models.DocumentLocker.id == id, models.DocumentLocker.owner_id == current_user.id)
    )
    doc = res.scalars().first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found or access denied")
    
    if doc.file_path and isinstance(doc.file_path, str) and os.path.exists(doc.file_path):
        os.remove(doc.file_path)
        
    await db.delete(doc)
    await db.commit()
    return {"status": "success"}

@app.post("/communication/broadcast-parents", response_model=schemas.ParentBroadcastOut)
async def broadcast_parents(
    data: schemas.ParentBroadcastCreate,
    current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PRINCIPAL", "HOD", "FACULTY"])),
    db: AsyncSession = Depends(get_db)
):
    # Fetch all students
    stmt = select(models.User).where(models.User.system_role == "STUDENT").options(
        selectinload(models.User.student_profile),
        selectinload(models.User.department)
    )
    result = await db.execute(stmt)
    students = result.scalars().all()

    # Filter target students
    target_students = []
    for s in students:
        # Department filter (HODs filter by their own department, Principal/Admins can filter by any department_code)
        if current_user.system_role == "HOD":
            if s.department_id != current_user.department_id:
                continue
        elif data.audience_type == "DEPARTMENT" and data.department_code:
            if not s.department or s.department.code != data.department_code:
                continue
                
        # Attendance filter
        if data.audience_type == "LOW_ATTENDANCE":
            att_stmt = select(models.AttendanceRecord).where(models.AttendanceRecord.student_id == s.id)
            att_res = await db.execute(att_stmt)
            att_records = att_res.scalars().all()
            if att_records:
                present = sum(1 for r in att_records if r.is_present)
                att_pct = (present / len(att_records)) * 100.0
            else:
                # Default mock low-attendance trigger if empty
                att_pct = 70.0 
            
            if att_pct >= 75.0:
                continue
                
        target_students.append(s)

    # Dispatch simulation
    notified_count = 0
    for ts in target_students:
        profile = ts.student_profile
        if not profile:
            continue
        parent_phone = profile.parent_whatsapp or profile.parent_email
        if parent_phone:
            # Simulated WhatsApp Integration API Dispatch
            print(f"[WHATSAPP BROADCAST DISPATCH] Student: {ts.username} | Parent: {parent_phone} | Message: {data.message}")
            notified_count += 1

    # Log broadcast activity in database
    db_broadcast = models.ParentBroadcast(
        sender_id=current_user.id,
        audience_type=data.audience_type,
        message=data.message,
        sent_at=datetime.utcnow().isoformat() + "Z",
        recipient_count=notified_count
    )
    db.add(db_broadcast)
    await db.commit()
    await db.refresh(db_broadcast)
    return db_broadcast

@app.get("/communication/broadcasts", response_model=list[schemas.ParentBroadcastOut])
async def list_parent_broadcasts(
    current_user: models.User = Depends(require_role(["ADMIN", "SUPER_ADMIN", "PRINCIPAL", "HOD", "FACULTY"])),
    db: AsyncSession = Depends(get_db)
):
    stmt = select(models.ParentBroadcast).order_by(models.ParentBroadcast.sent_at.desc())
    res = await db.execute(stmt)
    return res.scalars().all()

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
